import openpyxl
from openpyxl import load_workbook
excel_document = openpyxl.load_workbook("C:/Users/Fernando/testeufbot/tabula-2_ajuste_matriculas_2017.3_matriculas_por_aluno.xlsx")
sheet = excel_document.get_sheet_by_name('Worksheet')
def ra_find(RA): #Teste RA ALUNO
    for i in range(2, 42650):
    if RA = sheet.cell(row=i, column=1).value
        teste=sheet.cell(row=i, column=3).value
        teste.encode('utf-8')
        print(teste.encode('utf-8'))
