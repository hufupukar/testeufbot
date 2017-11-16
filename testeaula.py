import telepot
import openpyxl
from openpyxl import load_workbook
excel_document = openpyxl.load_workbook("C:/Users/Fernando/testeufbot/tabula-2_ajuste_matriculas_2017.3_matriculas_por_aluno.xlsx")
sheet = excel_document.get_sheet_by_name('Worksheet')


bot = telepot.Bot('474490928:AAEJVbHe0Rb5jx3sk1RDcqwbgMb8LnaQGHw')


def ModRecebeMsg(msg):
#modulo de processamendo de mensagens recebidas
	content_type, chat_type, chat_id = telepot.glance(msg)
	print(content_type, chat_type, chat_id)
	if content_type == "text":
		for i in range (2,42650):
			ra=sheet.cell(row=i, column=1).value
			if ra==11005316: #Aqui está onde não consigo saber o que fazer,
				teste=sheet.cell(row=i, column=3).value
				bot.sendMessage(chat_id, teste) #eu preciso que o RA seja igual ao texto inserido
                               #(supostamente o RA da pessoa)


bot.message_loop(ModRecebeMsg)
while True:
    pass


'''
#Se eu uso meu código da seguinte forma:

        def ModRecebeMsg(msg):
        #modulo de processamendo de mensagens recebidas
        	content_type, chat_type, chat_id = telepot.glance(msg)
        	print(content_type, chat_type, chat_id)
         	if content_type == 'text':
                    for i in range (2,42650):
                        ra=sheet.cell(row=i, column=1).value
                        bot.sendMessage(chat_id, ra)

        bot.message_loop(ModRecebeMsg)
        while True:
            pass

ele irá varrer todo o pdf e me mostrando todos os RA's, mas eu preciso que ele só me mostre quando
o RA for = ao inserido
'''
