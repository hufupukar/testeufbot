import openpyxl
from openpyxl import load_workbook
excel_document = openpyxl.load_workbook("C:/Users/Fernando/testeufbot/tabula-2_ajuste_matriculas_2017.3_matriculas_por_aluno.xlsx")
sheet = excel_document.get_sheet_by_name('Worksheet')

for i in range(2, 42650):
    value = sheet.cell(row=i, column=1).value
    if value==11005316:
            teste=sheet.cell(row=i, column=3).value
            teste.encode('utf-8')
            print(teste.encode('utf-8'))
